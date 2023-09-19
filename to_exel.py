import json
import openpyxl
def save_to_excel(file_1, file_2, file_3, file_output_excel):
    json_files = [file_1, file_2, file_3]
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet['A1'] = 'Название ссылки'
    worksheet['B1'] = 'Ссылка'
    row = 2
    for json_file in json_files:
        with open(json_file, 'r', encoding='utf-8') as file:
            data = json.load(file)
        for name, link in data.items():
            worksheet.cell(row=row, column=1, value=name)
            worksheet.cell(row=row, column=2, value=link)
            row += 1
    workbook.save(file_output_excel)
    print(f"Данные объединены и сохранены в {file_output_excel}")